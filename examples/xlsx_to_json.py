#!/usr/bin/env python
"""
把一张 Excel 表转成 `wp bangumi import` 用的 JSON。

这是一个「示例」脚本,按表头名称匹配列(不依赖列序):
  番名 -> title,评分 -> rating,评价/评语 -> review,观看方式 -> watch_mode
你的表头不同就改下面的 col(...) 名字,或直接手写 JSON(见 sample-anime.json)。

依赖: openpyxl  (pip install openpyxl)
用法: python xlsx_to_json.py <表格.xlsx> [输出.json]
"""
import json
import os
import sys

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("需要 openpyxl:pip install openpyxl")

if len(sys.argv) < 2:
    sys.exit("用法: python xlsx_to_json.py <表格.xlsx> [输出.json]")

SRC = sys.argv[1]
OUT = sys.argv[2] if len(sys.argv) > 2 else "anime.json"

wb = load_workbook(SRC, read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]
rows = list(ws.iter_rows(values_only=True))
if not rows:
    sys.exit("空表")

header = [str(c).strip() if c is not None else "" for c in rows[0]]


def col(*names):
    for i, h in enumerate(header):
        if h in names:
            return i
    return None


ci = {
    "title":  col("番名", "名称", "番剧", "title"),
    "rating": col("评分", "分数", "rating"),
    "review": col("评价", "评语", "短评", "review"),
    "note":   col("备注", "note"),
    "mode":   col("观看方式", "观看", "watch_mode"),
}
if ci["title"] is None:
    sys.exit(f"未找到番名列,表头={header}")


def get(row, i):
    if i is None or i >= len(row) or row[i] is None:
        return ""
    return str(row[i]).strip()


items = []
for row in rows[1:]:
    title = get(row, ci["title"])
    if not title:
        continue
    review = get(row, ci["review"])
    note = get(row, ci["note"])
    if note:  # 可选:把「备注」并进评语
        review = (review + "\n备注:" + note) if review else ("备注:" + note)
    items.append({
        "title":      title,
        "rating":     get(row, ci["rating"]),
        "review":     review,
        "watch_mode": get(row, ci["mode"]),
    })

os.makedirs(os.path.dirname(OUT) or ".", exist_ok=True)
with open(OUT, "w", encoding="utf-8") as f:
    json.dump(items, f, ensure_ascii=False, indent=2)
print(f"表头={header}")
print(f"写出 {len(items)} 条 -> {OUT}")
